#!/usr/bin/env python3
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries


SEGMENT_GROUPS = {
    "初中": ["初短一部", "初短二部", "初短三部", "郑州特战队"],
    "高中": ["小短", "高短"],
}
SEGMENTS = [seg for group in SEGMENT_GROUPS.values() for seg in group]
SEGMENT_KEYS = {
    "初短一部": "chuduan1",
    "初短二部": "chuduan2",
    "初短三部": "chuduan3",
    "郑州特战队": "tezhan",
    "小短": "xiaoduan",
    "高短": "gaoduan",
}
XINGHUO_TEAM_ALLOWLIST = {"星火先锋-秦智豪"}
TEAM_GRADE_OVERRIDE = {
    "溯川向上-刘炎鹤": "高二",
    "薪耀巅峰-李新": "高二",
}


def safe_div(a, b):
    try:
        a = float(a)
        b = float(b)
        if b == 0:
            return np.nan
        v = a / b
        return min(v, 1.0) if v == v else np.nan
    except Exception:
        return np.nan


def aggregate_auth_metrics(auth: pd.DataFrame, keys: List[str]) -> pd.DataFrame:
    """按老师邮箱去重后再汇总，避免重复明细行把授权人数重复累加。"""
    if auth.empty:
        return pd.DataFrame(columns=keys + ["授权人数", "正常人数"])
    work = auth.copy()
    work["老师邮箱"] = work["老师邮箱"].astype(str).str.strip().replace("nan", "")
    work = work[work["老师邮箱"] != ""]
    per_teacher = work.groupby(keys + ["老师邮箱"], as_index=False).agg(
        个微授权=("个微授权", "max"),
        整体正常=("整体正常", "max"),
    )
    return per_teacher.groupby(keys, as_index=False).agg(
        授权人数=("个微授权", "sum"),
        正常人数=("整体正常", "sum"),
    )


def cap_auth_counts(df: pd.DataFrame, flow_col: str = "接流") -> pd.DataFrame:
    out = df.copy()
    if flow_col not in out.columns:
        return out
    flow = pd.to_numeric(out[flow_col], errors="coerce").fillna(0)
    out["授权人数"] = pd.to_numeric(out.get("授权人数"), errors="coerce").fillna(0).clip(lower=0)
    out["正常人数"] = pd.to_numeric(out.get("正常人数"), errors="coerce").fillna(0).clip(lower=0)
    out["授权人数"] = np.minimum(out["授权人数"], flow)
    out["正常人数"] = np.minimum(out["正常人数"], out["授权人数"])
    return out


def dedupe_auth_detail(auth_detail: pd.DataFrame) -> pd.DataFrame:
    """同一业务日+战队+老师邮箱只保留一条，避免重复明细行重复累计授权人数。"""
    if auth_detail.empty:
        return auth_detail
    work = auth_detail.copy()
    work["老师邮箱"] = work["老师邮箱"].astype(str).str.strip().replace("nan", "")
    if "日期" in work.columns:
        work["日期"] = pd.to_datetime(work["日期"], errors="coerce").dt.strftime("%Y-%m-%d")
    keys = ["日期", "分组", "学部", "年级", "战队", "老师邮箱"]
    keys = [k for k in keys if k in work.columns]
    agg_map: Dict[str, str] = {}
    for col in work.columns:
        if col in keys:
            continue
        if col in {"个微授权", "整体正常"}:
            agg_map[col] = "max"
        else:
            agg_map[col] = "first"
    return work.groupby(keys, as_index=False).agg(agg_map)


def normalize_team_name(v):
    if pd.isna(v):
        return v
    s = str(v).strip()
    # 统一连接符，避免“--”和“-”导致同战队被拆成两条
    s = s.replace("——", "-")
    while "--" in s:
        s = s.replace("--", "-")
    return s


def remove_xinghuo_grade_rows(df: pd.DataFrame, grade_col: str = "年级") -> pd.DataFrame:
    if grade_col not in df.columns:
        return df
    keep_mask = ~df[grade_col].astype(str).str.contains("星火", na=False)
    return df[keep_mask].copy()


def remove_xinghuo_team_rows(df: pd.DataFrame, team_col: str = "战队") -> pd.DataFrame:
    if team_col not in df.columns:
        return df
    team_norm = df[team_col].astype(str).str.replace("——", "-", regex=False).str.replace("--", "-", regex=False)
    keep_mask = ~team_norm.str.contains("星火", na=False) | team_norm.isin(XINGHUO_TEAM_ALLOWLIST)
    return df[keep_mask].copy()


def apply_team_grade_override(df: pd.DataFrame, team_col: str = "战队", grade_col: str = "年级") -> pd.DataFrame:
    if team_col not in df.columns or grade_col not in df.columns:
        return df
    team_norm = df[team_col].astype(str).str.replace("——", "-", regex=False).str.replace("--", "-", regex=False)
    df = df.copy()
    for team_name, target_grade in TEAM_GRADE_OVERRIDE.items():
        df.loc[team_norm == team_name, grade_col] = target_grade
    return df


def drop_wrong_grade_team_rows(df: pd.DataFrame, team_col: str = "战队", grade_col: str = "年级") -> pd.DataFrame:
    """Remove duplicate team rows left on non-target grades after multi-source merges."""
    if df.empty or team_col not in df.columns or grade_col not in df.columns:
        return df
    data = df.copy()
    team_norm = data[team_col].astype(str).str.replace("——", "-", regex=False).str.replace("--", "-", regex=False)
    grade_norm = data[grade_col].astype(str).str.strip()
    keep = pd.Series(True, index=data.index)
    for team_name, target_grade in TEAM_GRADE_OVERRIDE.items():
        wrong = team_norm.eq(team_name) & grade_norm.ne(target_grade) & ~grade_norm.str.contains("汇总", na=False)
        keep &= ~wrong
    return data[keep].copy()


def norm_school(grade, src=None, xuebu=None):
    g = str(grade) if pd.notna(grade) else ""
    if g == "新兵营":
        if isinstance(src, str):
            if src == "高中":
                return "高中"
            if src == "爱学":
                return "初中"
        if isinstance(xuebu, str):
            if xuebu in ["高中", "初中"]:
                return xuebu
            if xuebu == "爱学":
                return "初中"
        return "初中"
    if g.startswith("小") or g == "全年级":
        return "小学"
    if g.startswith("高"):
        return "高中"
    if g.startswith("初"):
        return "初中"
    if isinstance(xuebu, str):
        if xuebu in ["高中", "初中"]:
            return xuebu
        if xuebu == "爱学":
            return "初中"
    return "其他"


def assign_segment(oc, src, grade=None):
    oc = "" if pd.isna(oc) else str(oc)
    src = "" if pd.isna(src) else str(src)
    g = "" if pd.isna(grade) else str(grade)
    # 高中源数据即使运营中心标成郑州一部，仍归高短（与授权口径一致）
    if src == "高中":
        return "高短"
    if oc in ("郑州三部", "郑州初短三部"):
        return "初短三部"
    if oc == "郑州特战队":
        return "郑州特战队"
    if oc == "郑州一部":
        return "初短一部"
    if oc == "郑州二部":
        return "初短二部"
    if oc == "郑州":
        # 新规则：郑州中心仅全年级归小短，其余郑州全部归高短
        if g == "全年级":
            return "小短"
        return "高短"
    return "其他"


def set_col_widths(ws, widths: Dict[int, int]):
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_raw_sheet(ws, df: pd.DataFrame):
    if df.empty:
        ws["A1"] = "无数据"
        return
    for j, c in enumerate(df.columns, 1):
        ws.cell(1, j, c)
    for i, row in enumerate(df.itertuples(index=False), 2):
        for j, v in enumerate(row, 1):
            ws.cell(i, j, None if pd.isna(v) else v)


def clone_cell_style(src_cell, tgt_cell):
    tgt_cell.font = copy(src_cell.font)
    tgt_cell.fill = copy(src_cell.fill)
    tgt_cell.border = copy(src_cell.border)
    tgt_cell.alignment = copy(src_cell.alignment)
    tgt_cell.number_format = src_cell.number_format
    tgt_cell.protection = copy(src_cell.protection)


def copy_conditional_formatting_from_template(template_ws, target_ws, data_start_row: int, max_row: int):
    target_ws.conditional_formatting = target_ws.conditional_formatting.__class__()
    cf_map = getattr(template_ws.conditional_formatting, "_cf_rules", {})
    for sqref, rules in cf_map.items():
        sqref_text = str(getattr(sqref, "sqref", sqref))
        new_ranges = []
        for ref in sqref_text.split():
            min_col, min_row, max_col, max_row_ref = range_boundaries(ref)
            if min_row >= data_start_row:
                end_row = max(max_row, min_row)
            else:
                end_row = max_row_ref
            new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{end_row}"
            new_ranges.append(new_ref)
        if not new_ranges:
            continue
        joined_ref = " ".join(new_ranges)
        for rule in rules:
            target_ws.conditional_formatting.add(joined_ref, copy(rule))


def apply_rate_font_colors(ws, start_row: int, end_row: int, rate_cols: List[int]):
    for r in range(start_row, end_row + 1):
        for c in rate_cols:
            cell = ws.cell(r, c)
            v = cell.value
            if v in (None, "", "#N/A"):
                continue
            try:
                x = float(v)
            except Exception:
                continue
            f = copy(cell.font)
            if x < 0.6:
                f.color = "C00000"
            elif x > 0.9:
                f.color = "00B050"
            else:
                f.color = "000000"
            cell.font = f


def clear_gray_fills(ws, start_row: int, end_row: int, max_col: int):
    gray_rgbs = {"FFD9D9D9", "00D9D9D9", "FFD0D0D0", "00D0D0D0"}
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
            if rgb in gray_rgbs:
                cell.fill = PatternFill(fill_type=None)


def pick_row(ws, start_row: int, pred):
    for r in range(start_row, ws.max_row + 1):
        if pred(r):
            return r
    return start_row


def collect_meta_auth(ws):
    detail = pick_row(ws, 3, lambda r: ws.cell(r, 3).value not in (None, "") and not (isinstance(ws.cell(r, 2).value, str) and "汇总" in ws.cell(r, 2).value))
    grade_sum = pick_row(ws, 3, lambda r: isinstance(ws.cell(r, 2).value, str) and "汇总" in ws.cell(r, 2).value and not (isinstance(ws.cell(r, 1).value, str) and "汇总" in ws.cell(r, 1).value))
    school_sum = pick_row(ws, 3, lambda r: isinstance(ws.cell(r, 1).value, str) and "汇总" in ws.cell(r, 1).value)
    return {
        "detail": detail,
        "grade_sum": grade_sum,
        "school_sum": school_sum,
        "col_widths": {c: ws.column_dimensions[get_column_letter(c)].width for c in range(1, 12)},
        "h1": ws.row_dimensions[1].height,
        "h2": ws.row_dimensions[2].height,
        "hd": ws.row_dimensions[detail].height,
        "hg": ws.row_dimensions[grade_sum].height,
        "hs": ws.row_dimensions[school_sum].height,
    }


def collect_meta_sales(ws):
    detail = pick_row(ws, 4, lambda r: ws.cell(r, 3).value not in (None, "") and not (isinstance(ws.cell(r, 2).value, str) and "汇总" in ws.cell(r, 2).value))
    grade_sum = pick_row(ws, 4, lambda r: isinstance(ws.cell(r, 2).value, str) and "汇总" in ws.cell(r, 2).value and not (isinstance(ws.cell(r, 1).value, str) and "汇总" in ws.cell(r, 1).value))
    school_sum = pick_row(ws, 4, lambda r: isinstance(ws.cell(r, 1).value, str) and "汇总" in ws.cell(r, 1).value)
    return {
        "detail": detail,
        "grade_sum": grade_sum,
        "school_sum": school_sum,
        "col_widths": {c: ws.column_dimensions[get_column_letter(c)].width for c in range(1, 9)},
        "h1": ws.row_dimensions[1].height,
        "h2": ws.row_dimensions[2].height,
        "h3": ws.row_dimensions[3].height,
        "hd": ws.row_dimensions[detail].height,
        "hg": ws.row_dimensions[grade_sum].height,
        "hs": ws.row_dimensions[school_sum].height,
    }


def rebuild_auth_public_sheet(ws, template_ws, meta, segment: str):
    for rg in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rg))

    maxr = ws.max_row
    ws.merge_cells("A1:K1")
    for c in range(1, 12):
        clone_cell_style(template_ws.cell(1, c), ws.cell(1, c))
        clone_cell_style(template_ws.cell(2, c), ws.cell(2, c))

    r = 3
    while r <= maxr:
        s_start = r
        s_end = r
        while s_end <= maxr:
            av = ws.cell(s_end, 1).value
            if isinstance(av, str) and "汇总" in av:
                break
            s_end += 1
        if s_end > maxr:
            s_end = maxr

        if s_start <= s_end - 1:
            ws.merge_cells(start_row=s_start, start_column=1, end_row=s_end - 1, end_column=1)

        rr = s_start
        while rr <= s_end - 1:
            bv = ws.cell(rr, 2).value
            if isinstance(bv, str) and "汇总" in bv:
                ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
                rr += 1
                continue
            g_start = rr
            g_end = rr
            while g_end + 1 <= s_end - 1:
                nb = ws.cell(g_end + 1, 2).value
                same_grade = (nb == bv) and not (isinstance(nb, str) and "汇总" in nb)
                both_blank = nb in (None, "") and bv in (None, "")
                if same_grade or both_blank:
                    g_end += 1
                else:
                    break
            if g_end > g_start:
                ws.merge_cells(start_row=g_start, start_column=2, end_row=g_end, end_column=2)
            rr = g_end + 1

        if s_end <= maxr and isinstance(ws.cell(s_end, 1).value, str) and "汇总" in ws.cell(s_end, 1).value:
            ws.merge_cells(start_row=s_end, start_column=1, end_row=s_end, end_column=3)
        r = s_end + 1

    for r in range(3, maxr + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        src_r = meta["detail"]
        is_summary = False
        if isinstance(a, str) and "汇总" in a:
            src_r = meta["school_sum"]
            ws.row_dimensions[r].height = meta["hs"]
            is_summary = True
        elif isinstance(b, str) and "汇总" in b:
            src_r = meta["grade_sum"]
            ws.row_dimensions[r].height = meta["hg"]
            is_summary = True
        else:
            ws.row_dimensions[r].height = meta["hd"]
        for c in range(1, 12):
            clone_cell_style(template_ws.cell(src_r, c), ws.cell(r, c))
            if is_summary:
                f = copy(ws.cell(r, c).font)
                f.bold = True
                ws.cell(r, c).font = f

    ws.row_dimensions[1].height = meta["h1"]
    ws.row_dimensions[2].height = meta["h2"]
    for c, w in meta["col_widths"].items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # 汇总行浅蓝底并保持加粗
    summary_fill = PatternFill("solid", fgColor="FFD9E5F3")
    for r in range(3, maxr + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        is_school_summary = isinstance(a, str) and "汇总" in a
        is_summary = is_school_summary or (isinstance(b, str) and "汇总" in b)
        if is_summary:
            for c in range(1, 12):
                cell = ws.cell(r, c)
                cell.fill = copy(summary_fill)
                f = copy(cell.font)
                f.bold = True
                cell.font = f

    # 汇总行后半段底色补齐：统一使用该行前段底色
    for r in range(3, maxr + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if (isinstance(a, str) and "汇总" in a) or (isinstance(b, str) and "汇总" in b):
            fill_src = ws.cell(r, 1).fill
            for c in range(1, 12):
                ws.cell(r, c).fill = copy(fill_src)

    # 数据条：按样式表2淡色系重设 F/J/K
    ws.conditional_formatting = ws.conditional_formatting.__class__()
    ws.conditional_formatting.add(
        f"F3:F{max(3, maxr)}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="FFBAC6F8", showValue=True),
    )
    ws.conditional_formatting.add(
        f"J3:J{max(3, maxr)}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="FFFCDA7B", showValue=True),
    )
    ws.conditional_formatting.add(
        f"K3:K{max(3, maxr)}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="FF7BE5BB", showValue=True),
    )

    # 在线率字体阈值规则：<60% 红，>90% 绿，其余黑
    apply_rate_font_colors(ws, 3, maxr, [6, 10, 11])
    for r in range(3, maxr + 1):
        for c in [6, 10, 11]:
            ws.cell(r, c).number_format = "0%"

    # 汇总行统一加粗（放在后面再兜底一次）
    for r in range(3, maxr + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if (isinstance(a, str) and "汇总" in a) or (isinstance(b, str) and "汇总" in b):
            for c in range(1, 12):
                f = copy(ws.cell(r, c).font)
                f.bold = True
                ws.cell(r, c).font = f

    # 学部汇总行：非在线率列橙色加粗（在线率列仍遵守阈值颜色规则）
    school_summary_rows = [r for r in range(3, maxr + 1) if isinstance(ws.cell(r, 1).value, str) and "汇总" in ws.cell(r, 1).value]
    for rr in school_summary_rows:
        for c in range(1, 12):
            f = copy(ws.cell(rr, c).font)
            f.bold = True
            if c not in [6, 10, 11]:
                f.color = "ED7D31"
            ws.cell(rr, c).font = f

    # 前两行统一居中加粗
    for r in [1, 2]:
        for c in range(1, 12):
            cell = ws.cell(r, c)
            f = copy(cell.font)
            f.bold = True
            cell.font = f
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 取消灰色底色
    clear_gray_fills(ws, 1, maxr, 11)


def rebuild_sales_display_sheet(ws, template_ws, meta):
    for rg in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rg))

    maxr = ws.max_row
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    ws.merge_cells("D2:D3")
    ws.merge_cells("E2:F2")
    ws.merge_cells("G2:H2")

    for c in range(1, 9):
        clone_cell_style(template_ws.cell(1, c), ws.cell(1, c))
        clone_cell_style(template_ws.cell(2, c), ws.cell(2, c))
        clone_cell_style(template_ws.cell(3, c), ws.cell(3, c))

    r = 4
    while r <= maxr:
        s_start = r
        s_end = r
        while s_end <= maxr:
            av = ws.cell(s_end, 1).value
            if isinstance(av, str) and "汇总" in av:
                break
            s_end += 1
        if s_end > maxr:
            s_end = maxr

        if s_start <= s_end - 1:
            ws.merge_cells(start_row=s_start, start_column=1, end_row=s_end - 1, end_column=1)

        rr = s_start
        while rr <= s_end - 1:
            bv = ws.cell(rr, 2).value
            if isinstance(bv, str) and "汇总" in bv:
                ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
                rr += 1
                continue
            g_start = rr
            g_end = rr
            while g_end + 1 <= s_end - 1:
                nb = ws.cell(g_end + 1, 2).value
                same_grade = (nb == bv) and not (isinstance(nb, str) and "汇总" in nb)
                both_blank = nb in (None, "") and bv in (None, "")
                if same_grade or both_blank:
                    g_end += 1
                else:
                    break
            if g_end > g_start:
                ws.merge_cells(start_row=g_start, start_column=2, end_row=g_end, end_column=2)
            rr = g_end + 1

        if s_end <= maxr and isinstance(ws.cell(s_end, 1).value, str) and "汇总" in ws.cell(s_end, 1).value:
            ws.merge_cells(start_row=s_end, start_column=1, end_row=s_end, end_column=3)
        r = s_end + 1

    for r in range(4, maxr + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        src_r = meta["detail"]
        if isinstance(a, str) and "汇总" in a:
            src_r = meta["school_sum"]
            ws.row_dimensions[r].height = meta["hs"]
        elif isinstance(b, str) and "汇总" in b:
            src_r = meta["grade_sum"]
            ws.row_dimensions[r].height = meta["hg"]
        else:
            ws.row_dimensions[r].height = meta["hd"]
        for c in range(1, 9):
            clone_cell_style(template_ws.cell(src_r, c), ws.cell(r, c))

    ws.row_dimensions[1].height = meta["h1"]
    ws.row_dimensions[2].height = meta["h2"]
    ws.row_dimensions[3].height = meta["h3"]
    for c, w in meta["col_widths"].items():
        ws.column_dimensions[get_column_letter(c)].width = w
    copy_conditional_formatting_from_template(template_ws, ws, data_start_row=4, max_row=max(4, maxr))
    for r in range(4, maxr + 1):
        for c in [6, 8]:
            ws.cell(r, c).number_format = "0%"
    # 表头样式严格按模板保持（避免后续逻辑二次覆盖）
    for c in range(1, 9):
        clone_cell_style(template_ws.cell(1, c), ws.cell(1, c))
        clone_cell_style(template_ws.cell(2, c), ws.cell(2, c))
        clone_cell_style(template_ws.cell(3, c), ws.cell(3, c))

    # 仅清理数据区灰底，不动表头
    clear_gray_fills(ws, 4, maxr, 8)


def apply_template_styles(output_root: Path, style_auth_template: Path, style_sales_template: Path, preserve_segment: str = "初短二部"):
    auth_wb = load_workbook(style_auth_template)
    sales_wb = load_workbook(style_sales_template)
    auth_tpl_ws = auth_wb["数据公示表"] if "数据公示表" in auth_wb.sheetnames else auth_wb[auth_wb.sheetnames[0]]
    sales_tpl_ws = sales_wb["战队数据"] if "战队数据" in sales_wb.sheetnames else sales_wb[sales_wb.sheetnames[0]]
    auth_meta = collect_meta_auth(auth_tpl_ws)
    sales_meta = collect_meta_sales(sales_tpl_ws)

    for seg in SEGMENTS:
        if preserve_segment and seg == preserve_segment:
            continue
        auth_fp = output_root / seg / f"郑州-{seg}-风灵个微全天在线率&爱芯后台授权.xlsx"
        sales_fp = output_root / seg / f"郑州-{seg}-销售风灵在线率明细数据.xlsx"
        if auth_fp.exists():
            awb = load_workbook(auth_fp)
            rebuild_auth_public_sheet(awb["数据公示表"], auth_tpl_ws, auth_meta, seg)
            awb.save(auth_fp)
        if sales_fp.exists():
            swb = load_workbook(sales_fp)
            rebuild_sales_display_sheet(swb["战队数据"], sales_tpl_ws, sales_meta)
            swb.save(sales_fp)


@dataclass
class DataBundle:
    sales: pd.DataFrame
    wechat_sum: pd.DataFrame
    wechat_detail: pd.DataFrame
    auth_detail: pd.DataFrame


SALES_DETAIL_SHEETS = ("销售风灵在线率明细数据", "学习规划师风灵在线率明细数据")
SOURCE_LIKE_XUEDUAN = {"高中", "爱学", "初中", "小学"}


def ensure_column(df: pd.DataFrame, target: str, *sources: str) -> pd.DataFrame:
    out = df.copy()
    if target in out.columns:
        return out
    for src in sources:
        if src in out.columns:
            out[target] = out[src]
            return out
    return out


def read_sales_detail(sales_path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(sales_path)
    for sheet in SALES_DETAIL_SHEETS:
        if sheet in xls.sheet_names:
            return pd.read_excel(sales_path, sheet_name=sheet)
    raise ValueError(
        f"销售源文件缺少明细 sheet，期望其一: {', '.join(SALES_DETAIL_SHEETS)}；实际: {xls.sheet_names}"
    )


def normalize_sales_export(df: pd.DataFrame) -> pd.DataFrame:
    out = ensure_column(df, "年级", "阶段")
    out = ensure_column(out, "老师姓名", "学习规划师姓名")
    out = ensure_column(out, "销售邮箱", "学习规划师邮箱")
    return out


def normalize_wechat_export(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "年级" not in out.columns and "阶段" in out.columns:
        out["年级"] = out["阶段"]
    if "学部" not in out.columns and "学段" in out.columns:
        xueduan_vals = set(out["学段"].dropna().astype(str).str.strip())
        if xueduan_vals & SOURCE_LIKE_XUEDUAN:
            out["学部"] = out["学段"]
    out = ensure_column(out, "辅导老师", "学习规划师")
    out = ensure_column(out, "辅导姓名", "学习规划师姓名", "学习规划师")
    out = ensure_column(out, "辅导在线人数", "学习规划师在线人数")
    out = ensure_column(out, "排班销售人数", "排班学习规划师人数")
    return out


def normalize_auth_export(df: pd.DataFrame, sys_source: str) -> pd.DataFrame:
    out = ensure_column(df, "年级", "阶段")
    if "学部" not in out.columns:
        out["学部"] = out["学段"] if "学段" in out.columns else sys_source
    out = ensure_column(out, "辅导老师邮箱", "学习规划师邮箱")
    out = ensure_column(out, "辅导名字", "学习规划师名字", "学习规划师姓名")
    out = ensure_column(out, "辅导邮箱", "学习规划师邮箱")
    out = ensure_column(out, "老师姓名", "学习规划师姓名", "辅导名字", "学习规划师名字")
    return out


def filter_bundle_by_date(bundle: DataBundle, as_of: str) -> DataBundle:
    target = pd.to_datetime(as_of).normalize()

    def on_date(df: pd.DataFrame, col: str = "日期") -> pd.DataFrame:
        if col not in df.columns or df.empty:
            return df.copy()
        mask = pd.to_datetime(df[col], errors="coerce").dt.normalize() == target
        return df.loc[mask].copy()

    return DataBundle(
        sales=on_date(bundle.sales),
        wechat_sum=on_date(bundle.wechat_sum),
        wechat_detail=on_date(bundle.wechat_detail),
        auth_detail=on_date(bundle.auth_detail),
    )


def load_and_prepare(
    sales_path: Path,
    auth_high_path: Path,
    wechat_path: Path,
    auth_aixue_path: Path,
) -> DataBundle:
    sales = normalize_sales_export(read_sales_detail(sales_path))
    sales = remove_xinghuo_grade_rows(sales, "年级")
    sales = apply_team_grade_override(sales, "战队", "年级")
    sales = remove_xinghuo_team_rows(sales, "战队")
    sales["战队"] = sales["战队"].map(normalize_team_name)
    if "sys_source" not in sales.columns and "系统来源" in sales.columns:
        sales["sys_source"] = sales["系统来源"]
    if "sys_source" not in sales.columns:
        sales["sys_source"] = sales.get("学部", "").astype(str)
    sales["sys_source"] = sales["sys_source"].astype(str)
    sales["学部"] = [norm_school(g, s, None) for g, s in zip(sales["年级"], sales["sys_source"])]
    sales["分组"] = [assign_segment(o, s, g) for o, s, g in zip(sales["运营中心"], sales["sys_source"], sales["年级"])]
    sales = sales[sales["分组"] != "其他"].copy()

    x_wechat = pd.ExcelFile(wechat_path)
    if "风灵微信在线率汇总数据" in x_wechat.sheet_names:
        wechat_sum = normalize_wechat_export(pd.read_excel(wechat_path, sheet_name="风灵微信在线率汇总数据"))
        wechat_sum = remove_xinghuo_grade_rows(wechat_sum, "年级")
        wechat_sum = apply_team_grade_override(wechat_sum, "战队", "年级")
        wechat_sum = remove_xinghuo_team_rows(wechat_sum, "战队")
        wechat_sum["战队"] = wechat_sum["战队"].map(normalize_team_name)
        wechat_sum["sys_source"] = wechat_sum["学部"].replace({"爱学": "爱学", "高中": "高中"})
        wechat_sum["学部"] = [
            norm_school(g, s, x) for g, s, x in zip(wechat_sum["年级"], wechat_sum["sys_source"], wechat_sum["学部"])
        ]
        wechat_sum["分组"] = [assign_segment(o, s, g) for o, s, g in zip(wechat_sum["运营中心"], wechat_sum["sys_source"], wechat_sum["年级"])]
        wechat_sum = wechat_sum[wechat_sum["分组"] != "其他"].copy()
    else:
        # 新版导出可能不含汇总sheet；保留空表占位，核心逻辑走明细聚合
        wechat_sum = pd.DataFrame(
            columns=["服务期", "日期", "分部", "学部", "年级", "运营中心", "战队", "辅导在线人数", "排班销售人数", "时段", "微信在线率", "sys_source", "分组"]
        )

    wechat_detail = normalize_wechat_export(pd.read_excel(wechat_path, sheet_name="风灵个微在线率明细数据"))
    wechat_detail = remove_xinghuo_grade_rows(wechat_detail, "年级")
    wechat_detail = apply_team_grade_override(wechat_detail, "战队", "年级")
    wechat_detail = remove_xinghuo_team_rows(wechat_detail, "战队")
    wechat_detail["战队"] = wechat_detail["战队"].map(normalize_team_name)
    wechat_detail["sys_source"] = wechat_detail["学部"].replace({"爱学": "爱学", "高中": "高中"})
    wechat_detail["学部"] = [
        norm_school(g, s, x) for g, s, x in zip(wechat_detail["年级"], wechat_detail["sys_source"], wechat_detail["学部"])
    ]
    wechat_detail["分组"] = [assign_segment(o, s, g) for o, s, g in zip(wechat_detail["运营中心"], wechat_detail["sys_source"], wechat_detail["年级"])]
    wechat_detail = wechat_detail[wechat_detail["分组"] != "其他"].copy()

    # 授权明细统一小时=23
    auth_h = normalize_auth_export(pd.read_excel(auth_high_path, sheet_name="个微授权明细数据"), "高中")
    auth_h = remove_xinghuo_grade_rows(auth_h, "年级")
    auth_h = apply_team_grade_override(auth_h, "战队", "年级")
    auth_h = remove_xinghuo_team_rows(auth_h, "战队")
    auth_h["战队"] = auth_h["战队"].map(normalize_team_name)
    auth_h["小时"] = pd.to_numeric(auth_h.get("小时"), errors="coerce")
    auth_h["sys_source"] = "高中"
    auth_h["老师邮箱"] = auth_h["辅导老师邮箱"]
    auth_h["老师姓名"] = auth_h["辅导名字"]
    auth_h["学部"] = [norm_school(g, "高中", x) for g, x in zip(auth_h["年级"], auth_h["学部"])]
    # 高中授权导出里运营中心可能标成郑州一部，但口径仍按郑州高短划分
    auth_h["分组"] = [assign_segment("郑州", "高中", g) for g in auth_h["年级"]]

    auth_a = normalize_auth_export(pd.read_excel(auth_aixue_path, sheet_name="个微授权明细数据"), "爱学")
    auth_a = remove_xinghuo_grade_rows(auth_a, "年级")
    auth_a = apply_team_grade_override(auth_a, "战队", "年级")
    auth_a = remove_xinghuo_team_rows(auth_a, "战队")
    auth_a["战队"] = auth_a["战队"].map(normalize_team_name)
    auth_a["小时"] = pd.to_numeric(auth_a.get("小时"), errors="coerce")
    auth_a["sys_source"] = "爱学"
    auth_a["老师邮箱"] = auth_a["辅导邮箱"]
    auth_a["老师姓名"] = auth_a["老师姓名"]
    auth_a["学部"] = [norm_school(g, "爱学", x) for g, x in zip(auth_a["年级"], auth_a["学部"])]
    auth_a["分组"] = [assign_segment(o, "爱学", g) for o, g in zip(auth_a["运营中心"], auth_a["年级"])]

    auth_cols = ["分组", "学部", "年级", "战队", "老师姓名", "老师邮箱", "个微授权", "整体正常", "运营中心", "日期", "小时"]
    auth_detail = pd.concat([auth_h[auth_cols], auth_a[auth_cols]], ignore_index=True)
    auth_detail = auth_detail[auth_detail["分组"] != "其他"].copy()
    for c in ["个微授权", "整体正常"]:
        auth_detail[c] = pd.to_numeric(auth_detail[c], errors="coerce").fillna(0)
    auth_detail = dedupe_auth_detail(auth_detail)

    return DataBundle(sales=sales, wechat_sum=wechat_sum, wechat_detail=wechat_detail, auth_detail=auth_detail)


def build_metrics(bundle: DataBundle, segment: str):
    sales = bundle.sales[bundle.sales["分组"] == segment].copy()
    wechat_sum = bundle.wechat_sum[bundle.wechat_sum["分组"] == segment].copy()
    wechat_detail = bundle.wechat_detail[bundle.wechat_detail["分组"] == segment].copy()
    auth = bundle.auth_detail[bundle.auth_detail["分组"] == segment].copy()

    # 业务约束：高短不允许出现“初中/小学”学部标签，统一归并到高短
    # 这样可避免高短中出现“初中 汇总”等中间汇总行导致口径错乱。
    if segment == "高短":
        for df in (sales, wechat_sum, wechat_detail, auth):
            if "学部" in df.columns:
                df["学部"] = "高短"

    keys = ["学部", "年级", "战队"]

    auth_teacher_count = auth.groupby(keys, as_index=False).agg(
        接流_auth=("老师邮箱", lambda s: s.astype(str).replace("nan", np.nan).dropna().nunique())
    )
    auth_metric = aggregate_auth_metrics(auth, keys)
    auth_agg = pd.merge(auth_teacher_count, auth_metric, on=keys, how="outer")

    # 个微透视：必须由“个微在线源数据”聚合得到
    wx_src = wechat_detail.copy()
    wx_src["老师标识"] = (
        wx_src.get("辅导老师", pd.Series(index=wx_src.index, dtype=object))
        .fillna(wx_src.get("辅导姓名", pd.Series(index=wx_src.index, dtype=object)))
        .astype(str)
    )
    wx_src["风灵微信在线率"] = pd.to_numeric(wx_src.get("风灵微信在线率"), errors="coerce")
    wx_src["老师标识"] = wx_src["老师标识"].replace("nan", np.nan)
    wx_teacher = wx_src.groupby(keys, as_index=False).agg(接流_wx=("老师标识", lambda s: s.dropna().nunique()))
    wx_pvt = wx_src.groupby(keys, as_index=False).agg(
        低价课带班=("老师标识", lambda s: s.replace("nan", np.nan).dropna().nunique()),
        全天在线=("风灵微信在线率", lambda s: int((s >= 1).sum())),
    )
    wx_pvt["个微全天在线率"] = [safe_div(a, b) for a, b in zip(wx_pvt["全天在线"], wx_pvt["低价课带班"])]

    sales_low = sales[sales["推广产品"].astype(str).str.contains("低价", na=False)].copy()
    sales_teacher = sales_low.groupby(keys, as_index=False).agg(
        接流_sales=(
            "销售邮箱",
            lambda s: s.astype(str).replace("nan", np.nan).dropna().nunique(),
        )
    )

    # 统一接流口径：多源去重人数的最大值，保证在线人数不会大于接流人数
    base = pd.merge(auth_teacher_count, wx_teacher, on=keys, how="outer")
    base = pd.merge(base, sales_teacher, on=keys, how="outer")
    for c in ["接流_auth", "接流_wx", "接流_sales"]:
        if c not in base.columns:
            base[c] = np.nan
    base["接流"] = base[["接流_auth", "接流_wx", "接流_sales"]].max(axis=1, skipna=True).fillna(0)
    base = base[keys + ["接流"]]

    func = pd.merge(base, wx_pvt, on=keys, how="left")
    func = pd.merge(func, auth_metric, on=keys, how="left")
    func["低价课带班"] = func["接流"]
    func["全天在线"] = func["全天在线"].fillna(0).clip(lower=0)
    func["全天在线"] = np.minimum(func["全天在线"], func["低价课带班"])
    func["授权人数"] = func["授权人数"].fillna(0).clip(lower=0)
    func["正常人数"] = func["正常人数"].fillna(0).clip(lower=0)
    func = cap_auth_counts(func, flow_col="低价课带班")
    func["个微全天在线率"] = [safe_div(a, b) for a, b in zip(func["全天在线"], func["低价课带班"])]
    func["爱芯个微授权率"] = [safe_div(a, b) for a, b in zip(func["授权人数"], func["低价课带班"])]
    func["个微功能正常率"] = [safe_div(a, b) for a, b in zip(func["正常人数"], func["授权人数"])]
    func = func[["学部", "年级", "战队", "低价课带班", "全天在线", "个微全天在线率", "接流", "授权人数", "正常人数", "爱芯个微授权率", "个微功能正常率"]]
    func = drop_wrong_grade_team_rows(func)

    # 企微在线人数按战队全量销售明细统计，避免接流含授权名单但在线只算低价课导致口径不一致
    online_counts = sales.groupby(keys, as_index=False).agg(
        电脑端全天在线人数=("pc在线率", lambda s: int((pd.to_numeric(s, errors="coerce") >= 1).sum())),
        手机端全天在线人数=("app在线率", lambda s: int((pd.to_numeric(s, errors="coerce") >= 1).sum())),
    )
    online = pd.merge(base.rename(columns={"接流": "接流人数"}), online_counts, on=keys, how="left")
    online["电脑端全天在线人数"] = online["电脑端全天在线人数"].fillna(0).clip(lower=0)
    online["手机端全天在线人数"] = online["手机端全天在线人数"].fillna(0).clip(lower=0)
    online["电脑端全天在线人数"] = np.minimum(online["电脑端全天在线人数"], online["接流人数"])
    online["手机端全天在线人数"] = np.minimum(online["手机端全天在线人数"], online["接流人数"])
    online["电脑端全天在线率"] = [safe_div(a, b) for a, b in zip(online["电脑端全天在线人数"], online["接流人数"])]
    online["手机端全天在线率"] = [safe_div(a, b) for a, b in zip(online["手机端全天在线人数"], online["接流人数"])]
    online = online[["学部", "年级", "战队", "接流人数", "电脑端全天在线人数", "电脑端全天在线率", "手机端全天在线人数", "手机端全天在线率"]]
    online = drop_wrong_grade_team_rows(online)

    bad_base = sales_low[["学部", "年级", "战队", "老师姓名", "app在线率", "pc在线率"]].copy()
    wx_d = wechat_detail[["学部", "年级", "战队", "辅导姓名", "风灵微信在线率"]].copy()
    wx_d = wx_d.rename(columns={"辅导姓名": "老师姓名", "风灵微信在线率": "个微在线率"})
    bad = pd.merge(bad_base, wx_d, on=["学部", "年级", "战队", "老师姓名"], how="left")
    bad = bad.rename(columns={"老师姓名": "辅导姓名", "app在线率": "企微-手机在线率", "pc在线率": "企微-电脑在线率"})
    for c in ["企微-手机在线率", "企微-电脑在线率", "个微在线率"]:
        bad[c] = pd.to_numeric(bad[c], errors="coerce")
    bad = bad[
        (bad["企微-手机在线率"] < 1) | (bad["企微-电脑在线率"] < 1) | (bad["个微在线率"] < 1)
    ].sort_values(["学部", "年级", "战队", "辅导姓名"])
    bad = bad[["学部", "年级", "战队", "辅导姓名", "企微-手机在线率", "企微-电脑在线率", "个微在线率"]]

    auth_export = pd.merge(base, auth_metric, on=keys, how="left")
    auth_export = cap_auth_counts(auth_export, flow_col="接流")
    return sales, wechat_sum, wechat_detail, auth, func, online, bad, wx_pvt, auth_export


def expand_with_summary(df: pd.DataFrame, metrics: List[str]) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    rows = []
    for xuebu, sdf in df.groupby("学部", dropna=False):
        for grade, gdf in sdf.groupby("年级", dropna=False):
            gdf = gdf.sort_values("战队")
            rows.append(gdf)
            line = {"学部": xuebu, "年级": f"{grade} 汇总", "战队": ""}
            for c in metrics:
                line[c] = gdf[c].sum(skipna=True)
            if "个微全天在线率" in df.columns:
                line["个微全天在线率"] = safe_div(line["全天在线"], line["低价课带班"])
                line["爱芯个微授权率"] = safe_div(line["授权人数"], line["低价课带班"])
                line["个微功能正常率"] = safe_div(line["正常人数"], line["授权人数"])
            if "电脑端全天在线率" in df.columns:
                line["电脑端全天在线率"] = safe_div(line["电脑端全天在线人数"], line["接流人数"])
                line["手机端全天在线率"] = safe_div(line["手机端全天在线人数"], line["接流人数"])
            rows.append(pd.DataFrame([line]))
        sline = {"学部": xuebu, "年级": f"{xuebu} 汇总", "战队": ""}
        for c in metrics:
            sline[c] = sdf[c].sum(skipna=True)
        if "个微全天在线率" in df.columns:
            sline["个微全天在线率"] = safe_div(sline["全天在线"], sline["低价课带班"])
            sline["爱芯个微授权率"] = safe_div(sline["授权人数"], sline["低价课带班"])
            sline["个微功能正常率"] = safe_div(sline["正常人数"], sline["授权人数"])
        if "电脑端全天在线率" in df.columns:
            sline["电脑端全天在线率"] = safe_div(sline["电脑端全天在线人数"], sline["接流人数"])
            sline["手机端全天在线率"] = safe_div(sline["手机端全天在线人数"], sline["接流人数"])
        rows.append(pd.DataFrame([sline]))
    return pd.concat(rows, ignore_index=True)


def write_public_sheet(ws, title: str, df: pd.DataFrame, pct_cols: List[str]):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_fill = PatternFill("solid", fgColor="EEE1F4")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    summary_fill = PatternFill("solid", fgColor="DCE6F2")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(color="7030A0", bold=True, size=16)
    ws.cell(1, 1).alignment = center
    ws.cell(1, 1).fill = title_fill

    for j, c in enumerate(df.columns, 1):
        x = ws.cell(2, j, c)
        x.fill = header_fill
        x.font = Font(color="FFFFFF", bold=True)
        x.alignment = center
        x.border = border

    for i, row in enumerate(df.itertuples(index=False), 3):
        summary = isinstance(row[df.columns.get_loc("年级")], str) and "汇总" in row[df.columns.get_loc("年级")]
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, "#N/A" if pd.isna(v) else v)
            c.border = border
            c.alignment = center
            if summary:
                c.fill = summary_fill
                c.font = Font(bold=True)
            if df.columns[j - 1] in pct_cols and c.value != "#N/A":
                c.number_format = "0.00%"
    ws.freeze_panes = "A3"


def write_online_sheet(ws, title: str, df: pd.DataFrame):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_fill = PatternFill("solid", fgColor="EEE1F4")
    blue = PatternFill("solid", fgColor="0F6DB3")
    orange = PatternFill("solid", fgColor="F79646")
    summary_fill = PatternFill("solid", fgColor="DCE6F2")

    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(color="7030A0", bold=True, size=16)
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    ws.merge_cells("D2:D3")
    ws.merge_cells("E2:F2")
    ws.merge_cells("G2:H2")
    ws["A2"], ws["B2"], ws["C2"], ws["D2"] = "学部", "年级", "战队", "接流人数"
    ws["E2"], ws["G2"] = "电脑端全天在线情况", "手机端全天在线情况"
    ws["E3"], ws["F3"], ws["G3"], ws["H3"] = "人数", "全天在线率", "人数", "全天在线率"

    for r in [2, 3]:
        for c in range(1, 9):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = center
            cell.fill = blue if c <= 6 else orange
            cell.font = Font(color="FFFFFF", bold=True)

    for i, row in enumerate(df.itertuples(index=False), 4):
        summary = isinstance(row[df.columns.get_loc("年级")], str) and "汇总" in row[df.columns.get_loc("年级")]
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, "#N/A" if pd.isna(v) else v)
            c.border = border
            c.alignment = center
            if summary:
                c.fill = summary_fill
                c.font = Font(bold=True)
            if j in [6, 8] and c.value != "#N/A":
                c.number_format = "0.00%"
    ws.freeze_panes = "A4"


def write_bad_sheet(ws, title: str, df: pd.DataFrame):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    title_fill = PatternFill("solid", fgColor="EEE1F4")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    bad_fill = PatternFill("solid", fgColor="F8CBAD")

    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = Font(color="7030A0", bold=True, size=16)
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    for j, c in enumerate(df.columns, 1):
        x = ws.cell(2, j, c)
        x.fill = header_fill
        x.font = Font(bold=True)
        x.alignment = center
        x.border = border

    for i, row in enumerate(df.itertuples(index=False), 3):
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, "#N/A" if pd.isna(v) else v)
            c.border = border
            c.alignment = left if j == 4 else center
            if j >= 5 and c.value != "#N/A":
                c.number_format = "0.00%"
                try:
                    if float(c.value) < 1:
                        c.fill = bad_fill
                        c.font = Font(color="C00000", bold=True)
                except Exception:
                    pass
    ws.freeze_panes = "A3"


def generate_segment_reports(bundle: DataBundle, segment: str, output_dir: Path, date_text: str):
    output_dir.mkdir(parents=True, exist_ok=True)
    sales, wechat_sum, wechat_detail, auth, func, online, bad, wx_pvt, auth_pvt = build_metrics(bundle, segment)

    func = expand_with_summary(
        func,
        metrics=["低价课带班", "全天在线", "接流", "授权人数", "正常人数"],
    )
    online = expand_with_summary(
        online,
        metrics=["接流人数", "电脑端全天在线人数", "手机端全天在线人数"],
    )

    # 文件1：个微+爱芯
    wb1 = Workbook()
    ws = wb1.active
    ws.title = "个微在线源数据"
    write_raw_sheet(ws, wechat_detail)
    ws2 = wb1.create_sheet("个微数据透视表")
    write_raw_sheet(ws2, wx_pvt)
    ws3 = wb1.create_sheet("数据公示表")
    fcols = [
        "学部",
        "年级",
        "战队",
        "低价课带班",
        "全天在线",
        "个微全天在线率",
        "接流",
        "授权人数",
        "正常人数",
        "爱芯个微授权率",
        "个微功能正常率",
    ]
    write_public_sheet(ws3, f"个微-风灵全天在线率及爱芯授权功能正常率（{segment}） {date_text}", func[fcols], ["个微全天在线率", "爱芯个微授权率", "个微功能正常率"])
    ws4 = wb1.create_sheet("爱芯后台授权原始数据")
    write_raw_sheet(ws4, auth)
    ws5 = wb1.create_sheet("爱芯透视")
    auth_pvt_export = auth_pvt.copy()
    auth_pvt_export["爱芯个微授权率"] = [safe_div(a, b) for a, b in zip(auth_pvt_export["授权人数"], auth_pvt_export["接流"])]
    auth_pvt_export["个微功能正常率"] = [safe_div(a, b) for a, b in zip(auth_pvt_export["正常人数"], auth_pvt_export["授权人数"])]
    write_raw_sheet(ws5, auth_pvt_export)
    set_col_widths(ws3, {1: 8, 2: 10, 3: 24, 4: 11, 5: 10, 6: 14, 7: 9, 8: 10, 9: 10, 10: 14, 11: 14})
    wb1.save(output_dir / f"郑州-{segment}-风灵个微全天在线率&爱芯后台授权.xlsx")

    # 文件2：销售在线
    wb2 = Workbook()
    x = wb2.active
    x.title = "销售风灵在线率明细数据"
    write_raw_sheet(x, sales)
    x2 = wb2.create_sheet("战队汇总透视")
    write_raw_sheet(x2, online)
    x3 = wb2.create_sheet("战队数据")
    ocols = ["学部", "年级", "战队", "接流人数", "电脑端全天在线人数", "电脑端全天在线率", "手机端全天在线人数", "手机端全天在线率"]
    write_online_sheet(x3, f"企微风灵全天在线情况通晒（{segment}）-{date_text}", online[ocols])
    set_col_widths(x3, {1: 8, 2: 10, 3: 24, 4: 10, 5: 12, 6: 14, 7: 12, 8: 14})
    wb2.save(output_dir / f"郑州-{segment}-销售风灵在线率明细数据.xlsx")

    # 文件3：未达标
    wb3 = Workbook()
    y = wb3.active
    y.title = "原表"
    write_raw_sheet(y, sales[sales["推广产品"].astype(str).str.contains("低价", na=False)])
    y2 = wb3.create_sheet("Sheet11")
    write_raw_sheet(y2, wechat_detail)
    y3 = wb3.create_sheet("数据透视表")
    bcols = ["学部", "年级", "战队", "辅导姓名", "企微-手机在线率", "企微-电脑在线率", "个微在线率"]
    write_bad_sheet(y3, f"风灵在线未达标名单（{segment}）-{date_text}（低价课）", bad[bcols])
    set_col_widths(y3, {1: 8, 2: 10, 3: 24, 4: 12, 5: 14, 6: 14, 7: 12})
    wb3.save(output_dir / f"郑州-{segment}-每日风灵不在线.xlsx")


def main():
    parser = argparse.ArgumentParser(description="根据4个原始文件生成四个分组日报文件夹")
    parser.add_argument("--sales", required=True, help="销售风灵在线率明细数据xlsx路径")
    parser.add_argument("--auth-high", required=True, help="爱芯个微授权数据（高中）xlsx路径")
    parser.add_argument("--wechat", required=True, help="风灵个微在线数据xlsx路径")
    parser.add_argument("--auth-aixue", required=True, help="爱芯个微授权数据（爱学）xlsx路径")
    parser.add_argument("--output-root", required=True, help="输出根目录")
    parser.add_argument("--style-auth-template", default="", help="样式模板：风灵个微文件路径（默认用输出目录下初短二部文件）")
    parser.add_argument("--style-sales-template", default="", help="样式模板：销售风灵文件路径（默认用输出目录下初短二部文件）")
    parser.add_argument("--as-of-date", default="", help="仅处理指定业务日期（YYYY-MM-DD），用于周维度历史回填")
    args = parser.parse_args()

    bundle = load_and_prepare(
        sales_path=Path(args.sales),
        auth_high_path=Path(args.auth_high),
        wechat_path=Path(args.wechat),
        auth_aixue_path=Path(args.auth_aixue),
    )
    if args.as_of_date:
        bundle = filter_bundle_by_date(bundle, args.as_of_date)
        if bundle.sales.empty and bundle.auth_detail.empty and bundle.wechat_detail.empty:
            raise SystemExit(f"指定日期无任何数据: {args.as_of_date}")
        if bundle.sales.empty:
            print(f"警告：{args.as_of_date} 无销售数据，将仅基于授权/个微数据生成。")

    dt = pd.to_datetime(bundle.sales.get("日期"), errors="coerce").max()
    date_text = f"{dt.month}月{dt.day}日" if pd.notna(dt) else ""

    output_root = Path(args.output_root)
    output_root.mkdir(parents=True, exist_ok=True)
    for seg in SEGMENTS:
        generate_segment_reports(bundle, seg, output_root / seg, date_text)

    # 高短数据检阅：不得出现“初中”学部字段
    _, _, _, _, high_func, high_online, _, _, _ = build_metrics(bundle, "高短")
    school_values = set(pd.concat([high_func["学部"], high_online["学部"]], axis=0).dropna().astype(str).tolist())
    if "初中" in school_values:
        print("警告：高短检阅未通过，仍存在初中字段。", school_values)
    else:
        print("高短检阅通过：未出现初中字段。")

    default_auth_tpl = output_root / "初短二部" / "郑州-初短二部-风灵个微全天在线率&爱芯后台授权.xlsx"
    default_sales_tpl = output_root / "初短二部" / "郑州-初短二部-销售风灵在线率明细数据.xlsx"
    auth_tpl = Path(args.style_auth_template) if args.style_auth_template else default_auth_tpl
    sales_tpl = Path(args.style_sales_template) if args.style_sales_template else default_sales_tpl
    if auth_tpl.exists() and sales_tpl.exists():
        apply_template_styles(output_root, auth_tpl, sales_tpl, preserve_segment="初短二部")
    else:
        print("警告：样式模板文件不存在，跳过套版。")

    print("完成输出目录：", output_root)
    print("分组：", ", ".join(SEGMENTS))


if __name__ == "__main__":
    main()
